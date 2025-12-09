# -*- coding: utf-8 -*-
"""
组织信息管理蓝图

该文件负责处理学院、系、专业等组织信息的通用查询API，
提供可被其他蓝图共享的组织信息查询功能。
"""

from flask import Blueprint, request, jsonify
import os
import uuid
import openpyxl
from werkzeug.utils import secure_filename
from models import Faculty, Department, Major

# 创建蓝图实例
organization_bp = Blueprint("organization", __name__, url_prefix="/api/organization")


# 通用辅助函数：将模型对象转换为字典
def model_to_dict(model, detailed=False, extra_fields=None):
    """
    将模型对象转换为字典

    Args:
        model: 数据库模型对象
        detailed: 是否返回详细信息
        extra_fields: 额外需要添加的字段字典

    Returns:
        包含模型信息的字典
    """
    # 基本信息
    result = {"id": model.id, "name": model.name}

    # 详细信息（如果需要）
    if (
        detailed
        and hasattr(model, "description")
        and hasattr(model, "created_at")
        and hasattr(model, "updated_at")
    ):
        result.update(
            {
                "description": model.description,
                "created_at": model.created_at.isoformat(),
                "updated_at": model.updated_at.isoformat(),
            }
        )

    # 添加额外字段
    if extra_fields:
        result.update(extra_fields)

    return result


# 通用函数：获取学院列表
def get_all_faculties(detailed=False):
    """
    获取所有学院信息

    Args:
        detailed: 是否返回详细信息（包含描述、创建时间等）

    Returns:
        学院信息列表
    """
    faculties = Faculty.query.all()
    return [model_to_dict(faculty, detailed) for faculty in faculties]


# 通用函数：根据学院ID获取系列表
def get_departments_by_faculty_id(faculty_id, detailed=False):
    """
    根据学院ID获取系列表

    Args:
        faculty_id: 学院ID
        detailed: 是否返回详细信息

    Returns:
        系列表
    """
    departments = Department.query.filter_by(faculty_id=faculty_id).all()

    result = []
    for dept in departments:
        extra_fields = {"faculty_id": dept.faculty_id} if detailed else None
        result.append(model_to_dict(dept, detailed, extra_fields))

    return result


# 通用函数：获取所有系列表
def get_all_departments(detailed=False):
    """
    获取所有系信息列表

    Args:
        detailed: 是否返回详细信息

    Returns:
        系列表
    """
    departments = Department.query.all()

    if not detailed:
        return [model_to_dict(dept, detailed) for dept in departments]

    # 预加载所有相关的学院信息以避免N+1查询
    faculty_ids = {dept.faculty_id for dept in departments}
    faculties = {
        f.id: f for f in Faculty.query.filter(Faculty.id.in_(faculty_ids)).all()
    }

    result = []
    for dept in departments:
        faculty = faculties.get(dept.faculty_id)
        extra_fields = {
            "faculty_id": dept.faculty_id,
            "faculty_name": faculty.name if faculty else "未知学院",
        }
        result.append(model_to_dict(dept, detailed, extra_fields))

    return result


# 通用函数：获取所有专业列表
def get_all_majors(detailed=False):
    """
    获取所有专业列表

    Args:
        detailed: 是否返回详细信息

    Returns:
        专业列表
    """
    majors = Major.query.all()

    if not detailed:
        return [model_to_dict(major, detailed) for major in majors]

    # 预加载所有相关的系和学院信息以避免N+1查询
    major_ids = [major.id for major in majors]
    department_ids = {major.department_id for major in majors}

    departments = {
        d.id: d
        for d in Department.query.filter(Department.id.in_(department_ids)).all()
    }
    faculty_ids = {d.faculty_id for d in departments.values()}
    faculties = {
        f.id: f for f in Faculty.query.filter(Faculty.id.in_(faculty_ids)).all()
    }

    result = []
    for major in majors:
        department = departments.get(major.department_id)
        faculty = faculties.get(department.faculty_id) if department else None

        extra_fields = {
            "department_id": major.department_id,
            "department_name": department.name if department else "未知系",
            "faculty_id": faculty.id if faculty else None,
            "faculty_name": faculty.name if faculty else "未知学院",
        }

        result.append(model_to_dict(major, detailed, extra_fields))

    return result


# 通用函数：根据学院ID获取专业列表
def get_majors_by_faculty_id(faculty_id, detailed=False):
    """
    根据学院ID获取专业列表

    Args:
        faculty_id: 学院ID
        detailed: 是否返回详细信息

    Returns:
        专业列表
    """
    # 通过系表关联查询学院下的所有专业
    majors = (
        Major.query.join(Department).filter(Department.faculty_id == faculty_id).all()
    )

    if not detailed:
        return [model_to_dict(major, detailed) for major in majors]

    # 预加载所有相关的系和学院信息以避免N+1查询
    department_ids = {major.department_id for major in majors}
    departments = {
        d.id: d
        for d in Department.query.filter(Department.id.in_(department_ids)).all()
    }

    # 预加载学院信息
    faculty = Faculty.query.filter_by(id=faculty_id).first()
    faculty_name = faculty.name if faculty else "未知学院"

    result = []
    for major in majors:
        department = departments.get(major.department_id)
        extra_fields = {
            "department_id": major.department_id,
            "department_name": department.name if department else "未知系",
            "faculty_id": faculty_id,
            "faculty_name": faculty_name,
        }
        result.append(model_to_dict(major, detailed, extra_fields))

    return result


# 通用函数：根据系ID获取专业列表
def get_majors_by_department_id(department_id, detailed=False):
    """
    根据系ID获取专业列表

    Args:
        department_id: 系ID
        detailed: 是否返回详细信息

    Returns:
        专业列表
    """
    majors = Major.query.filter_by(department_id=department_id).all()

    if not detailed:
        return [model_to_dict(major, detailed) for major in majors]

    # 预加载所有相关的系和学院信息以避免N+1查询
    departments = {
        d.id: d for d in Department.query.filter(Department.id == department_id).all()
    }
    faculty_ids = {d.faculty_id for d in departments.values()}
    faculties = {
        f.id: f for f in Faculty.query.filter(Faculty.id.in_(faculty_ids)).all()
    }

    result = []
    for major in majors:
        department = departments.get(major.department_id)
        faculty = faculties.get(department.faculty_id) if department else None

        extra_fields = {
            "department_id": major.department_id,
            "department_name": department.name if department else "未知系",
            "faculty_id": faculty.id if faculty else None,
            "faculty_name": faculty.name if faculty else "未知学院",
        }
        result.append(model_to_dict(major, detailed, extra_fields))

    return result


# API端点：获取所有学院
@organization_bp.route("/faculties", methods=["GET"])
def get_faculties():
    """
    获取所有学院信息的API端点
    可选参数：detailed=true 返回详细信息
    """
    detailed = request.args.get("detailed", "false").lower() == "true"
    faculties = get_all_faculties(detailed)
    return jsonify({"faculties": faculties}), 200


# API端点：根据学院ID获取系列表
@organization_bp.route("/departments/<int:faculty_id>", methods=["GET"])
def get_departments(faculty_id):
    """
    根据学院ID获取系列表的API端点
    可选参数：detailed=true 返回详细信息
    """
    detailed = request.args.get("detailed", "false").lower() == "true"
    departments = get_departments_by_faculty_id(faculty_id, detailed)
    return jsonify({"departments": departments}), 200


# API端点：获取所有专业列表
@organization_bp.route("/majors", methods=["GET"])
def get_majors():
    """
    获取所有专业列表的API端点
    可选参数：detailed=true 返回详细信息
    """
    detailed = request.args.get("detailed", "false").lower() == "true"
    majors = get_all_majors(detailed)
    return jsonify({"majors": majors}), 200


# API端点：根据系ID获取专业列表
@organization_bp.route("/majors/department/<int:department_id>", methods=["GET"])
def get_majors_by_department(department_id):
    """
    根据系ID获取专业列表的API端点
    可选参数：detailed=true 返回详细信息
    """
    detailed = request.args.get("detailed", "false").lower() == "true"
    majors = get_majors_by_department_id(department_id, detailed)
    return jsonify({"majors": majors}), 200


# 导入组织数据
@organization_bp.route("/import-organizations", methods=["POST"])
def import_organizations():
    try:
        # 检查是否有文件上传
        if "file" not in request.files:
            return jsonify({"message": "没有文件上传"}), 400

        file = request.files["file"]

        # 检查文件名是否为空
        if file.filename == "":
            return jsonify({"message": "请选择一个文件"}), 400

        # 检查文件类型
        allowed_extensions = {".xlsx", ".xls"}
        _, ext = os.path.splitext(file.filename.lower())
        if ext not in allowed_extensions:
            return jsonify({"message": "只支持Excel文件(.xlsx, .xls)"}), 400

        # 读取Excel文件
        workbook = openpyxl.load_workbook(file, data_only=True)

        # 定义各表的数据
        faculty_data = []
        department_data = []
        major_data = []

        # 单表导入模式
        target_sheet = workbook[workbook.sheetnames[0]]

        # 解析表头
        headers = {}
        for col in range(1, target_sheet.max_column + 1):
            header = target_sheet.cell(row=1, column=col).value
            if header:
                headers[header.strip()] = col

        # 验证必填表头（至少需要学院名称）
        required_headers = ["学院名称"]
        missing_headers = [h for h in required_headers if h not in headers]
        if missing_headers:
            return (
                jsonify(
                    {"message": f"工作表缺少必填表头：{','.join(missing_headers)}"}
                ),
                400,
            )

        # 读取数据行
        for row in range(2, target_sheet.max_row + 1):
            # 学院信息
            faculty_name = target_sheet.cell(row=row, column=headers["学院名称"]).value
            faculty_description = (
                target_sheet.cell(row=row, column=headers.get("学院描述")).value
                if "学院描述" in headers
                else None
            )

            # 系信息
            department_name = (
                target_sheet.cell(row=row, column=headers.get("系名称")).value
                if "系名称" in headers
                else None
            )
            department_description = (
                target_sheet.cell(row=row, column=headers.get("系描述")).value
                if "系描述" in headers
                else None
            )

            # 专业信息
            major_name = (
                target_sheet.cell(row=row, column=headers.get("专业名称")).value
                if "专业名称" in headers
                else None
            )
            major_description = (
                target_sheet.cell(row=row, column=headers.get("专业描述")).value
                if "专业描述" in headers
                else None
            )

            # 处理学院
            if faculty_name:
                faculty_data.append(
                    {
                        "name": faculty_name.strip(),
                        "description": (
                            faculty_description.strip() if faculty_description else None
                        ),
                    }
                )

            # 处理系
            if faculty_name and department_name:
                department_data.append(
                    {
                        "name": department_name.strip(),
                        "faculty_name": faculty_name.strip(),
                        "description": (
                            department_description.strip()
                            if department_description
                            else None
                        ),
                    }
                )

            # 处理专业
            if faculty_name and department_name and major_name:
                major_data.append(
                    {
                        "name": major_name.strip(),
                        "faculty_name": faculty_name.strip(),
                        "department_name": department_name.strip(),
                        "description": (
                            major_description.strip() if major_description else None
                        ),
                    }
                )

        # 去重数据
        # 学院去重
        unique_faculties = {}
        for data in faculty_data:
            unique_faculties[data["name"]] = data
        faculty_data = list(unique_faculties.values())

        # 系去重
        unique_departments = {}
        for data in department_data:
            key = (data["faculty_name"], data["name"])
            unique_departments[key] = data
        department_data = list(unique_departments.values())

        # 专业去重
        unique_majors = {}
        for data in major_data:
            key = (data["faculty_name"], data["department_name"], data["name"])
            unique_majors[key] = data
        major_data = list(unique_majors.values())

        # 导入数据到数据库
        from extensions import db

        errors = []

        # 导入学院
        faculty_map = {}
        for data in faculty_data:
            # 检查学院是否已存在
            existing_faculty = Faculty.query.filter_by(name=data["name"]).first()
            if not existing_faculty:
                faculty = Faculty(name=data["name"], description=data["description"])
                db.session.add(faculty)
                db.session.flush()  # 获取新插入的ID
                faculty_map[data["name"]] = faculty.id
            else:
                # 更新已存在学院的描述信息
                if (
                    data["description"]
                    and existing_faculty.description != data["description"]
                ):
                    existing_faculty.description = data["description"]
                    errors.append(f"学院 '{data['name']}' 的描述已更新")
                faculty_map[data["name"]] = existing_faculty.id
                if (
                    not data["description"]
                    or existing_faculty.description == data["description"]
                ):
                    errors.append(f"学院 '{data['name']}' 已存在，未重复导入")

        # 导入系
        department_map = {}
        for data in department_data:
            if data["faculty_name"] in faculty_map:
                # 检查系是否已存在
                existing_department = Department.query.filter_by(
                    name=data["name"], faculty_id=faculty_map[data["faculty_name"]]
                ).first()
                if not existing_department:
                    department = Department(
                        name=data["name"],
                        faculty_id=faculty_map[data["faculty_name"]],
                        description=data["description"],
                    )
                    db.session.add(department)
                    db.session.flush()  # 获取新插入的ID
                    department_map[(data["faculty_name"], data["name"])] = department.id
                else:
                    # 更新已存在系的描述信息
                    if (
                        data["description"]
                        and existing_department.description != data["description"]
                    ):
                        existing_department.description = data["description"]
                        errors.append(
                            f"系 '{data['name']}' 在学院 '{data['faculty_name']}' 下的描述已更新"
                        )
                    department_map[(data["faculty_name"], data["name"])] = (
                        existing_department.id
                    )
                    if (
                        not data["description"]
                        or existing_department.description == data["description"]
                    ):
                        errors.append(
                            f"系 '{data['name']}' 在学院 '{data['faculty_name']}' 下已存在，未重复导入"
                        )
            else:
                errors.append(
                    f"系 '{data['name']}' 的所属学院 '{data['faculty_name']}' 不存在，未导入"
                )

        # 导入专业
        for data in major_data:
            key = (data["faculty_name"], data["department_name"])
            if data["faculty_name"] in faculty_map and key in department_map:
                # 检查专业是否已存在
                existing_major = Major.query.filter_by(
                    name=data["name"], department_id=department_map[key]
                ).first()
                if not existing_major:
                    major = Major(
                        name=data["name"],
                        department_id=department_map[key],
                        description=data["description"],
                    )
                    db.session.add(major)
                else:
                    # 更新已存在专业的描述信息
                    if (
                        data["description"]
                        and existing_major.description != data["description"]
                    ):
                        existing_major.description = data["description"]
                        errors.append(
                            f"专业 '{data['name']}' 在学院 '{data['faculty_name']}'、系 '{data['department_name']}' 下的描述已更新"
                        )
                    if (
                        not data["description"]
                        or existing_major.description == data["description"]
                    ):
                        errors.append(
                            f"专业 '{data['name']}' 在学院 '{data['faculty_name']}'、系 '{data['department_name']}' 下已存在，未重复导入"
                        )
            else:
                errors.append(
                    f"专业 '{data['name']}' 的所属学院 '{data['faculty_name']}' 或系 '{data['department_name']}' 不存在，未导入"
                )

        # 提交事务
        db.session.commit()

        return (
            jsonify(
                {
                    "message": "导入成功",
                    "faculty_count": len(faculty_data),
                    "department_count": len(department_data),
                    "major_count": len(major_data),
                    "errors": errors,
                }
            ),
            200,
        )

    except Exception as e:
        return jsonify({"message": str(e)}), 500
