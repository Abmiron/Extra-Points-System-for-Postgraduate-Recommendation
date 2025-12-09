# -*- coding: utf-8 -*-
"""
用户管理蓝图

该文件负责处理用户管理相关的API端点，包括：
- 获取用户信息
- 管理员获取所有用户
- 管理员删除用户
- 管理员更新用户信息
- 管理员重置用户密码
- 用户上传头像
- 学生信息管理（增删改查）
"""

from flask import Blueprint, request, jsonify, current_app, session
from models import User, Faculty, Department, Major, Student, Application
import os
import uuid
from PIL import Image
from io import BytesIO
from werkzeug.utils import secure_filename
from extensions import db
import openpyxl
from openpyxl.utils import get_column_letter

# 导入组织信息管理模块，用于获取学院、系、专业信息
from blueprints.organization_bp import (
    get_all_faculties,
    get_all_departments,
    get_all_majors,
)

# 创建蓝图实例
user_bp = Blueprint("user", __name__, url_prefix="/api")


# 获取单个用户信息接口
@user_bp.route("/user/<username>", methods=["GET"])
def get_user(username):
    user = User.query.filter_by(username=username).first()

    if not user:
        return jsonify({"message": "用户不存在"}), 404

    return _get_user_data(user)


# 获取当前登录用户信息接口
@user_bp.route("/user/current", methods=["GET"])
def get_current_user():
    # 从session获取用户名，提高安全性
    if (
        "username" not in session
        or "logged_in" not in session
        or not session["logged_in"]
    ):
        return jsonify({"message": "未登录或会话已过期"}), 401

    username = session["username"]
    user = User.query.filter_by(username=username).first()

    if not user:
        return jsonify({"message": "用户不存在"}), 404

    return _get_user_data(user)


# 获取当前学生用户信息接口（为了解决前端调用/api/user/student的问题）
@user_bp.route("/user/student", methods=["GET"])
def get_student_user():
    # 从session获取用户名，提高安全性
    if (
        "username" not in session
        or "logged_in" not in session
        or not session["logged_in"]
    ):
        return jsonify({"message": "未登录或会话已过期"}), 401

    username = session["username"]
    user = User.query.filter_by(username=username, role="student").first()

    if not user:
        return jsonify({"message": "学生用户不存在"}), 404

    return _get_user_data(user)


# 辅助函数：获取用户的组织信息
def _get_organization_info(user):
    """
    获取用户的学院、系和专业信息

    Args:
        user: 用户对象

    Returns:
        包含学院、系和专业名称的字典
    """
    # 预加载所有组织信息，避免重复查询
    all_faculties = {f["id"]: f["name"] for f in get_all_faculties()}
    all_departments = {d["id"]: d["name"] for d in get_all_departments()}
    all_majors = {m["id"]: m["name"] for m in get_all_majors()}

    # 安全获取学院、系和专业名称
    faculty_name = all_faculties.get(user.faculty_id, "")
    department_name = all_departments.get(user.department_id, "")
    major_name = all_majors.get(user.major_id, "")

    return {
        "faculty_name": faculty_name,
        "department_name": department_name,
        "major_name": major_name,
    }


# 辅助函数：获取用户数据
def _get_user_data(user):
    # 获取用户的组织信息
    org_info = _get_organization_info(user)

    # 返回用户信息（不包含密码）
    user_data = {
        "id": user.id,
        "username": user.username,
        "name": user.name,
        "role": user.role,
        "avatar": user.avatar,
        "faculty": org_info["faculty_name"],
        "facultyId": user.faculty_id,
        "department": org_info["department_name"],
        "departmentId": user.department_id,
        "major": org_info["major_name"],
        "majorId": user.major_id,
        "studentId": user.student_id,
        "email": user.email,
        "phone": user.phone,
        "roleName": user.role_name,
        "lastLogin": user.last_login.isoformat() if user.last_login else None,
    }

    return jsonify({"user": user_data}), 200


# 辅助函数：检查用户是否为管理员
def is_admin(user):
    return user and user.role == "admin"


# 管理员获取所有用户接口
@user_bp.route("/admin/users", methods=["GET"])
def get_all_users():
    # 获取请求参数
    role = request.args.get("role")
    status = request.args.get("status")
    search = request.args.get("search")
    faculty = request.args.get("faculty")
    department = request.args.get("department")
    major = request.args.get("major")
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 10, type=int)

    # 调试日志
    print(
        f"筛选参数: role={role}, status={status}, faculty={faculty}, department={department}, major={major}, search={search}"
    )

    # 构建查询
    query = User.query

    # 根据role筛选
    if role:
        query = query.filter_by(role=role)

    # 根据status筛选
    if status:
        query = query.filter_by(status=status)

    # 根据学院ID筛选
    if faculty and faculty != "all":
        try:
            faculty_id = int(faculty)
            query = query.filter_by(faculty_id=faculty_id)
        except ValueError:
            # 如果不是有效的ID，尝试根据学院名称查询
            faculty_obj = Faculty.query.filter_by(name=faculty).first()
            if faculty_obj:
                query = query.filter_by(faculty_id=faculty_obj.id)

    # 根据系ID筛选
    if department and department != "all":
        try:
            department_id = int(department)
            query = query.filter_by(department_id=department_id)
        except ValueError:
            # 如果不是有效的ID，尝试根据系名称查询
            department_obj = Department.query.filter_by(name=department).first()
            if department_obj:
                query = query.filter_by(department_id=department_obj.id)

    # 根据专业ID筛选
    if major and major != "all":
        try:
            major_id = int(major)
            query = query.filter_by(major_id=major_id)
        except ValueError:
            # 如果不是有效的ID，尝试根据专业名称查询
            major_obj = Major.query.filter_by(name=major).first()
            if major_obj:
                query = query.filter_by(major_id=major_obj.id)

    # 根据search关键词搜索
    if search:
        query = query.filter(
            (User.name.like(f"%{search}%")) | (User.username.like(f"%{search}%"))
        )

    # 按账号（username）排序
    query = query.order_by(User.username)

    # 分页查询
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    users = pagination.items

    # 获取所有用户列表
    user_list = []
    # 一次性获取所有组织信息，避免循环中的重复查询
    all_faculties = {f["id"]: f["name"] for f in get_all_faculties()}
    all_departments = {d["id"]: d["name"] for d in get_all_departments()}
    all_majors = {m["id"]: m["name"] for m in get_all_majors()}

    for user in users:
        # 从预加载的数据中获取组织信息
        faculty_name = all_faculties.get(user.faculty_id, "")
        department_name = all_departments.get(user.department_id, "")
        major_name = all_majors.get(user.major_id, "")

        user_data = {
            "id": user.id,
            "username": user.username,
            "name": user.name,
            "role": user.role,
            "avatar": user.avatar,
            "faculty": faculty_name,
            "facultyId": user.faculty_id,
            "department": department_name,
            "departmentId": user.department_id,
            "major": major_name,
            "majorId": user.major_id,
            "studentId": user.student_id,
            "email": user.email,
            "phone": user.phone,
            "roleName": user.role_name,
            "status": user.status,
            "lastLogin": user.last_login.isoformat() if user.last_login else None,
        }
        user_list.append(user_data)

    # 返回分页数据
    return (
        jsonify(
            {
                "users": user_list,
                "total": pagination.total,
                "pages": pagination.pages,
                "current_page": pagination.page,
            }
        ),
        200,
    )


# 管理员删除用户接口
@user_bp.route("/admin/users/<int:user_id>", methods=["DELETE"])
def delete_user(user_id):
    # 获取当前登录用户的ID
    current_user_id = request.args.get("currentUserId", type=int)

    # 检查是否尝试删除自己
    if current_user_id == user_id:
        return jsonify({"message": "无法删除自己的账号"}), 400

    user = User.query.get(user_id)

    if not user:
        return jsonify({"message": "用户不存在"}), 404

    # 如果是学生用户，同时删除关联的Student记录和申请数据
    if user.role == "student" and user.student:
        # 获取该学生的所有申请数据
        applications = Application.query.filter_by(
            student_id=user.student.student_id
        ).all()
        # 先删除每个申请关联的文件
        for application in applications:
            if application.files:
                for file in application.files:
                    try:
                        if "path" in file and file["path"]:
                            filename = os.path.basename(file["path"])
                            file_path = os.path.join(
                                current_app.config["FILE_FOLDER"], filename
                            )
                            if os.path.exists(file_path):
                                os.remove(file_path)
                    except Exception as e:
                        print(f"删除文件失败: {str(e)}")
        # 删除该学生的所有申请数据
        Application.query.filter_by(student_id=user.student.student_id).delete()
        # 删除Student记录
        db.session.delete(user.student)

    db.session.delete(user)
    db.session.commit()

    return jsonify({"message": "用户删除成功"}), 200


# 管理员更新用户信息接口
@user_bp.route("/admin/users/<int:user_id>", methods=["PUT"])
def update_user(user_id):
    # 获取当前登录用户的ID
    current_user_id = request.args.get("currentUserId", type=int)

    # 检查是否尝试修改自己
    if current_user_id == user_id:
        return jsonify({"message": "无法修改自己的账号信息"}), 400

    user = User.query.get(user_id)

    if not user:
        return jsonify({"message": "用户不存在"}), 404

    data = request.get_json()

    # 更新用户信息
    if "name" in data:
        user.name = data["name"]
    if "role" in data:
        user.role = data["role"]
    if "facultyId" in data:
        user.faculty_id = data["facultyId"]
    if "departmentId" in data:
        user.department_id = data["departmentId"]
    if "majorId" in data:
        user.major_id = data["majorId"]
    if "email" in data:
        user.email = data["email"]
    if "phone" in data:
        user.phone = data["phone"]
    if "roleName" in data:
        user.role_name = data["roleName"]
    if "status" in data:
        user.status = data["status"]

    db.session.commit()

    return jsonify({"message": "用户信息更新成功"}), 200


# 管理员添加用户接口
@user_bp.route("/admin/create-users", methods=["POST", "OPTIONS"])
def create_user():
    # 处理OPTIONS预检请求
    if request.method == "OPTIONS":
        return jsonify({}), 200
    # 获取当前登录用户的ID
    current_user_id = request.args.get("currentUserId", type=int)

    # 验证请求数据
    data = request.get_json()
    if not data:
        return jsonify({"message": "请求数据不能为空"}), 400

    # 显式删除可能存在的student_id字段，避免使用前端传递的值
    if "student_id" in data:
        del data["student_id"]

    # 验证必填字段
    required_fields = ["username", "name", "role", "status"]
    for field in required_fields:
        if field not in data or not data[field]:
            return jsonify({"message": f"{field}是必填字段"}), 400

    # 检查用户名是否已存在
    existing_user = User.query.filter_by(username=data["username"]).first()
    if existing_user:
        return jsonify({"message": "用户名已存在"}), 400

    # 获取关联ID
    faculty_id = data.get("facultyId")
    department_id = data.get("departmentId")
    major_id = data.get("majorId")

    # 学生角色验证
    if data["role"] == "student":
        # 验证学生必填的关联字段
        if not faculty_id or not department_id or not major_id:
            return jsonify({"message": "学生用户必须选择学院、系和专业"}), 400

        # 验证关联ID是否存在
        faculty = Faculty.query.get(faculty_id)
        department = Department.query.get(department_id)
        major = Major.query.get(major_id)

        if not faculty:
            return jsonify({"message": "学院不存在"}), 400
        if not department:
            return jsonify({"message": "系不存在"}), 400
        if not major:
            return jsonify({"message": "专业不存在"}), 400

    try:
        # 学生角色需要先处理Student记录
        student_id = None
        if data["role"] == "student":
            # 检查是否已存在对应的Student记录
            existing_student = Student.query.filter_by(
                student_id=data["username"]
            ).first()
            if not existing_student:
                # 如果不存在，创建新的Student记录
                new_student = Student(
                    student_id=data["username"],
                    student_name=data["name"],  # 使用用户姓名作为学生姓名
                    faculty_id=faculty_id,
                    department_id=department_id,
                    major_id=major_id,
                )
                db.session.add(new_student)
                db.session.flush()  # 确保学生记录已创建
            student_id = data[
                "username"
            ]  # 使用username作为student_id，与注册接口保持一致

        # 创建用户记录，确保不使用前端传递的student_id
        new_user = User(
            username=data["username"],
            name=data["name"],
            role=data["role"],
            faculty_id=faculty_id,
            department_id=department_id,
            major_id=major_id,
            student_id=student_id,  # 只使用后端生成的student ID
            role_name=data.get(
                "roleName", "审核员" if data["role"] == "teacher" else "系统管理员"
            ),
            status=data["status"],
            email=data.get("email", ""),
            phone=data.get("phone", ""),
        )

        # 设置密码
        password = data.get("password", "123456")
        new_user.set_password(password)

        db.session.add(new_user)
        db.session.commit()

        return jsonify({"message": "用户创建成功"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"message": f"用户创建失败: {str(e)}"}), 500


# 管理员重置用户密码接口
@user_bp.route("/admin/users/<int:user_id>/reset-password", methods=["POST"])
def admin_reset_password(user_id):
    # 获取当前登录用户的ID
    current_user_id = request.args.get("currentUserId", type=int)

    # 检查是否尝试重置自己的密码
    if current_user_id == user_id:
        return jsonify({"message": "无法重置自己的密码，请使用个人设置修改密码"}), 400

    user = User.query.get(user_id)

    if not user:
        return jsonify({"message": "用户不存在"}), 404

    data = request.get_json()
    new_password = data.get("newPassword")

    if not new_password:
        return jsonify({"message": "新密码不能为空"}), 400

    # 设置新密码
    user.set_password(new_password)

    db.session.commit()

    return jsonify({"message": "密码重置成功"}), 200


# 用户更新个人信息接口
@user_bp.route("/user/profile", methods=["PUT"])
def update_profile():
    # 从session获取用户名，提高安全性
    if (
        "username" not in session
        or "logged_in" not in session
        or not session["logged_in"]
    ):
        return jsonify({"message": "未登录或会话已过期"}), 401

    username = session["username"]
    user = User.query.filter_by(username=username).first()

    if not user:
        return jsonify({"message": "用户不存在"}), 404

    # 获取请求数据
    data = request.get_json() or {}

    # 更新用户信息
    if "name" in data:
        user.name = data["name"]
    if "facultyId" in data:
        user.faculty_id = data["facultyId"]
    if "departmentId" in data:
        user.department_id = data["departmentId"]
    if "majorId" in data:
        user.major_id = data["majorId"]
    if "email" in data:
        user.email = data["email"]
    if "phone" in data:
        user.phone = data["phone"]

    db.session.commit()

    # 获取学院、系和专业名称
    faculty_name = Faculty.query.get(user.faculty_id).name if user.faculty_id else ""
    department_name = (
        Department.query.get(user.department_id).name if user.department_id else ""
    )
    major_name = Major.query.get(user.major_id).name if user.major_id else ""

    # 返回更新后的用户信息
    user_data = {
        "id": user.id,
        "username": user.username,
        "name": user.name,
        "role": user.role,
        "avatar": user.avatar,
        "faculty": faculty_name,
        "facultyId": user.faculty_id,
        "department": department_name,
        "departmentId": user.department_id,
        "major": major_name,
        "majorId": user.major_id,
        "studentId": user.student_id,
        "email": user.email,
        "phone": user.phone,
        "roleName": user.role_name,
        "lastLogin": user.last_login.isoformat() if user.last_login else None,
    }

    return jsonify({"user": user_data, "message": "个人信息更新成功"}), 200


# 恢复默认头像接口
@user_bp.route("/user/avatar/reset", methods=["POST"])
def reset_avatar():
    # 从session获取用户名，提高安全性
    if (
        "username" not in session
        or "logged_in" not in session
        or not session["logged_in"]
    ):
        return jsonify({"message": "未登录或会话已过期"}), 401

    username = session["username"]
    user = User.query.filter_by(username=username).first()

    if not user:
        return jsonify({"message": "用户不存在"}), 404

    try:
        # 将用户头像设置为空字符串，表示使用默认头像
        user.avatar = ""

        db.session.commit()

        # 获取学院、系和专业名称
        faculty_name = (
            Faculty.query.get(user.faculty_id).name if user.faculty_id else ""
        )
        department_name = (
            Department.query.get(user.department_id).name if user.department_id else ""
        )
        major_name = Major.query.get(user.major_id).name if user.major_id else ""

        # 返回更新后的用户信息
        user_data = {
            "id": user.id,
            "username": user.username,
            "name": user.name,
            "role": user.role,
            "avatar": user.avatar,
            "faculty": faculty_name,
            "facultyId": user.faculty_id,
            "department": department_name,
            "departmentId": user.department_id,
            "major": major_name,
            "majorId": user.major_id,
            "studentId": user.student_id,
            "email": user.email,
            "phone": user.phone,
            "roleName": user.role_name,
            "lastLogin": user.last_login.isoformat() if user.last_login else None,
        }

        return jsonify({"user": user_data, "message": "恢复默认头像成功"}), 200
    except Exception as e:
        return jsonify({"message": f"恢复默认头像失败: {str(e)}"}), 500


# 用户上传头像接口
@user_bp.route("/user/avatar", methods=["POST"])
def upload_avatar():
    # 从session获取用户名，提高安全性
    if (
        "username" not in session
        or "logged_in" not in session
        or not session["logged_in"]
    ):
        return jsonify({"message": "未登录或会话已过期"}), 401

    username = session["username"]
    user = User.query.filter_by(username=username).first()

    if not user:
        return jsonify({"message": "用户不存在"}), 404

    # 检查是否有文件上传
    if "avatar" not in request.files:
        return jsonify({"message": "没有文件上传"}), 400

    file = request.files["avatar"]

    # 检查文件名是否为空
    if file.filename == "":
        return jsonify({"message": "请选择一个文件"}), 400

    # 从系统设置中获取头像文件大小限制
    from models import SystemSettings

    settings = SystemSettings.query.first()
    avatar_file_size_limit = (
        settings.avatar_file_size_limit if settings else 2
    )  # 默认2MB
    max_size = avatar_file_size_limit * 1024 * 1024
    if request.content_length > max_size:
        return jsonify({"message": f"文件大小不能超过{avatar_file_size_limit}MB"}), 400

    # 增强文件类型验证
    # 检查文件扩展名
    allowed_extensions = {".png", ".jpg", ".jpeg", ".gif"}
    _, ext = os.path.splitext(file.filename.lower())
    if ext not in allowed_extensions:
        return jsonify({"message": "只支持PNG、JPG、JPEG和GIF格式的图片"}), 400

    # 检查MIME类型
    if not file.mimetype.startswith("image/"):
        return jsonify({"message": "上传的文件不是有效的图片"}), 400

    if not user:
        return jsonify({"message": "用户不存在"}), 404

    try:
        # 确保上传目录存在
        if not os.path.exists(current_app.config["UPLOAD_FOLDER"]):
            os.makedirs(current_app.config["UPLOAD_FOLDER"])

        # 确保头像上传目录存在
        if not os.path.exists(current_app.config["AVATAR_FOLDER"]):
            os.makedirs(current_app.config["AVATAR_FOLDER"])

        # 使用secure_filename防止路径遍历攻击，并生成唯一的文件名
        safe_filename = secure_filename(file.filename)
        _, ext = os.path.splitext(safe_filename)
        unique_filename = f"{uuid.uuid4()}{ext}"
        file_path = os.path.join(current_app.config["AVATAR_FOLDER"], unique_filename)

        # 处理图片：调整大小和压缩
        image = Image.open(file)

        # 最大尺寸限制
        max_size = (500, 500)

        # 调整图片大小
        image.thumbnail(max_size)

        # 创建字节流来保存处理后的图片
        output = BytesIO()

        # 保存图片（根据文件类型选择格式）
        if safe_filename.lower().endswith((".png", ".gif")):
            image.save(output, format="PNG", optimize=True)
        else:
            # JPEG格式可以压缩质量
            image.save(output, format="JPEG", quality=85, optimize=True)

        # 将字节流内容保存到文件
        with open(file_path, "wb") as f:
            f.write(output.getvalue())

        # 更新用户头像路径
        user.avatar = f"/uploads/avatars/{unique_filename}"

        db.session.commit()

        # 获取学院、系和专业名称
        faculty_name = (
            Faculty.query.get(user.faculty_id).name if user.faculty_id else ""
        )
        department_name = (
            Department.query.get(user.department_id).name if user.department_id else ""
        )
        major_name = Major.query.get(user.major_id).name if user.major_id else ""

        # 返回更新后的用户信息
        user_data = {
            "id": user.id,
            "username": user.username,
            "name": user.name,
            "role": user.role,
            "avatar": user.avatar,
            "faculty": faculty_name,
            "facultyId": user.faculty_id,
            "department": department_name,
            "departmentId": user.department_id,
            "major": major_name,
            "majorId": user.major_id,
            "studentId": user.student_id,
            "email": user.email,
            "phone": user.phone,
            "roleName": user.role_name,
            "lastLogin": user.last_login.isoformat() if user.last_login else None,
        }

        return jsonify({"user": user_data, "message": "头像上传成功"}), 200
    except Exception as e:
        return jsonify({"message": f"头像上传失败: {str(e)}"}), 500


# 用户修改密码接口
@user_bp.route("/user/change-password", methods=["POST"])
def change_password():
    # 从session获取用户名，提高安全性
    if (
        "username" not in session
        or "logged_in" not in session
        or not session["logged_in"]
    ):
        return jsonify({"message": "未登录或会话已过期"}), 401

    username = session["username"]
    user = User.query.filter_by(username=username).first()

    data = request.get_json()
    current_password = data.get("currentPassword")
    new_password = data.get("newPassword")

    if not current_password or not new_password:
        return jsonify({"message": "缺少必要参数"}), 400

    if not user:
        return jsonify({"message": "用户不存在"}), 404

    # 验证当前密码
    if not user.check_password(current_password):
        return jsonify({"message": "当前密码错误"}), 400

    # 设置新密码
    user.set_password(new_password)

    db.session.commit()

    return jsonify({"message": "密码修改成功"}), 200


# 管理员导入用户接口
@user_bp.route("/admin/import-users", methods=["POST"])
def import_users():
    # 检查用户是否为管理员
    if "role" not in session or session["role"] != "admin":
        return jsonify({"message": "权限不足"}), 403

    # 检查是否有文件上传
    if "file" not in request.files:
        return jsonify({"message": "没有文件上传"}), 400

    file = request.files["file"]

    # 检查文件名是否为空
    if file.filename == "":
        return jsonify({"message": "请选择一个文件"}), 400

    # 检查文件类型
    allowed_extensions = {".xlsx", ".xls"}
    _, ext = os.path.splitext(file.filename.lower())
    if ext not in allowed_extensions:
        return jsonify({"message": "只支持Excel文件(.xlsx, .xls)"}), 400

    try:
        # 读取Excel文件
        workbook = openpyxl.load_workbook(file, data_only=True)
        worksheet = workbook.active

        # 定义必填字段和可选字段
        required_headers = ["用户名", "姓名", "角色", "状态"]
        optional_headers = [
            "学院",
            "系",
            "专业",
            "邮箱",
            "电话",
            "性别",
            "CET4成绩",
            "CET6成绩",
            "绩点",
            "转换分数",
        ]

        # 读取表头并创建映射
        header_to_column = {}
        max_column = worksheet.max_column

        for col in range(1, max_column + 1):
            header = worksheet.cell(row=1, column=col).value
            if header:
                header_to_column[header.strip()] = col

        # 验证必填字段是否存在
        missing_headers = []
        for header in required_headers:
            if header not in header_to_column:
                missing_headers.append(header)

        if missing_headers:
            return (
                jsonify({"message": f"缺少必填表头：{','.join(missing_headers)}"}),
                400,
            )

        # 导入用户数据
        success_count = 0
        error_count = 0
        errors = []

        for row in range(2, worksheet.max_row + 1):
            try:
                # 根据表头映射获取用户数据
                username = worksheet.cell(
                    row=row, column=header_to_column["用户名"]
                ).value
                name = worksheet.cell(row=row, column=header_to_column["姓名"]).value
                role = worksheet.cell(row=row, column=header_to_column["角色"]).value
                status = worksheet.cell(row=row, column=header_to_column["状态"]).value

                # 可选字段
                faculty_name = (
                    worksheet.cell(row=row, column=header_to_column.get("学院")).value
                    if "学院" in header_to_column
                    else None
                )
                department_name = (
                    worksheet.cell(row=row, column=header_to_column.get("系")).value
                    if "系" in header_to_column
                    else None
                )
                major_name = (
                    worksheet.cell(row=row, column=header_to_column.get("专业")).value
                    if "专业" in header_to_column
                    else None
                )
                email = (
                    worksheet.cell(row=row, column=header_to_column.get("邮箱")).value
                    if "邮箱" in header_to_column
                    else None
                )
                phone = (
                    worksheet.cell(row=row, column=header_to_column.get("电话")).value
                    if "电话" in header_to_column
                    else None
                )
                gender = (
                    worksheet.cell(row=row, column=header_to_column.get("性别")).value
                    if "性别" in header_to_column
                    else None
                )

                # CET4成绩转换为整数
                cet4_score = (
                    worksheet.cell(
                        row=row, column=header_to_column.get("CET4成绩")
                    ).value
                    if "CET4成绩" in header_to_column
                    else None
                )
                if cet4_score is not None and cet4_score != "":
                    try:
                        cet4_score = int(cet4_score)
                    except ValueError:
                        cet4_score = None
                else:
                    cet4_score = None

                # CET6成绩转换为整数
                cet6_score = (
                    worksheet.cell(
                        row=row, column=header_to_column.get("CET6成绩")
                    ).value
                    if "CET6成绩" in header_to_column
                    else None
                )
                if cet6_score is not None and cet6_score != "":
                    try:
                        cet6_score = int(cet6_score)
                    except ValueError:
                        cet6_score = None
                else:
                    cet6_score = None

                # 绩点转换为浮点数
                gpa = (
                    worksheet.cell(row=row, column=header_to_column.get("绩点")).value
                    if "绩点" in header_to_column
                    else None
                )
                if gpa is not None and gpa != "":
                    try:
                        gpa = float(gpa)
                    except ValueError:
                        gpa = None
                else:
                    gpa = None

                # 转换分数转换为浮点数
                academic_score = (
                    worksheet.cell(
                        row=row, column=header_to_column.get("转换分数")
                    ).value
                    if "转换分数" in header_to_column
                    else None
                )
                if academic_score is not None and academic_score != "":
                    try:
                        academic_score = float(academic_score)
                    except ValueError:
                        academic_score = None
                else:
                    academic_score = None

                # 验证必填字段
                if not username or not name or not role or not status:
                    error_count += 1
                    errors.append(f"第{row}行：用户名、姓名、角色和状态为必填字段")
                    continue

                # 检查用户名是否已存在
                existing_user = User.query.filter_by(username=username).first()
                if existing_user:
                    # 获取或创建学院、系、专业ID（如果提供了名称）
                    update_faculty_id = None
                    update_department_id = None
                    update_major_id = None

                    # 处理学院ID
                    if faculty_name:
                        faculty = Faculty.query.filter_by(name=faculty_name).first()
                        if not faculty:
                            error_count += 1
                            errors.append(f"第{row}行：学院 '{faculty_name}' 不存在")
                            continue
                        update_faculty_id = faculty.id
                    elif faculty_name == "":
                        # 如果明确为空字符串，则清空该字段
                        update_faculty_id = None
                    else:
                        # 保留原有值
                        update_faculty_id = existing_user.faculty_id

                    # 处理系ID
                    if department_name:
                        department = Department.query.filter_by(
                            name=department_name
                        ).first()
                        if not department:
                            error_count += 1
                            errors.append(f"第{row}行：系 '{department_name}' 不存在")
                            continue
                        update_department_id = department.id
                    elif department_name == "":
                        # 如果明确为空字符串，则清空该字段
                        update_department_id = None
                    else:
                        # 保留原有值
                        update_department_id = existing_user.department_id

                    # 处理专业ID
                    if major_name:
                        major = Major.query.filter_by(name=major_name).first()
                        if not major:
                            error_count += 1
                            errors.append(f"第{row}行：专业 '{major_name}' 不存在")
                            continue
                        update_major_id = major.id
                    elif major_name == "":
                        # 如果明确为空字符串，则清空该字段
                        update_major_id = None
                    else:
                        # 保留原有值
                        update_major_id = existing_user.major_id

                    # 更新用户基本信息（适用于所有角色）
                    existing_user.name = name
                    existing_user.faculty_id = update_faculty_id
                    existing_user.department_id = update_department_id
                    existing_user.major_id = update_major_id
                    existing_user.email = email or ""
                    existing_user.phone = phone or ""

                    # 将enabled转换为active以与数据库模型保持一致
                    if status == "enabled":
                        status = "active"
                    existing_user.status = status

                    # 如果是学生用户，还需要更新学生特定信息
                    if existing_user.role == "student":
                        # 获取或创建学生记录
                        existing_student = Student.query.filter_by(
                            student_id=username
                        ).first()
                        if existing_student:
                            # 更新学生信息
                            existing_student.gender = gender
                            existing_student.cet4_score = cet4_score
                            existing_student.cet6_score = cet6_score
                            existing_student.gpa = gpa
                            existing_student.academic_score = academic_score
                        else:
                            # 如果学生记录不存在，创建新的学生记录
                            new_student = Student(
                                student_id=username,
                                student_name=name,
                                faculty_id=update_faculty_id,
                                department_id=update_department_id,
                                major_id=update_major_id,
                                gender=gender,
                                cet4_score=cet4_score,
                                cet6_score=cet6_score,
                                gpa=gpa,
                                academic_score=academic_score,
                            )
                            db.session.add(new_student)
                            existing_user.student_id = username

                    success_count += 1
                    continue

                # 验证角色
                valid_roles = ["admin", "teacher", "student"]
                if role not in valid_roles:
                    error_count += 1
                    errors.append(
                        f"第{row}行：角色 '{role}' 无效，应为 admin/teacher/student"
                    )
                    continue

                # 验证状态并统一转换为数据库使用的值
                valid_statuses = ["enabled", "disabled", "active"]
                if status not in valid_statuses:
                    error_count += 1
                    errors.append(
                        f"第{row}行：状态 '{status}' 无效，应为 enabled/disabled/active"
                    )
                    continue

                # 将enabled转换为active以与数据库模型保持一致
                if status == "enabled":
                    status = "active"

                # 获取学院ID
                faculty_id = None
                if faculty_name:
                    faculty = Faculty.query.filter_by(name=faculty_name).first()
                    if not faculty:
                        error_count += 1
                        errors.append(f"第{row}行：学院 '{faculty_name}' 不存在")
                        continue
                    faculty_id = faculty.id

                # 获取系ID
                department_id = None
                if department_name:
                    department = Department.query.filter_by(
                        name=department_name
                    ).first()
                    if not department:
                        error_count += 1
                        errors.append(f"第{row}行：系 '{department_name}' 不存在")
                        continue
                    department_id = department.id

                # 获取专业ID
                major_id = None
                if major_name:
                    major = Major.query.filter_by(name=major_name).first()
                    if not major:
                        error_count += 1
                        errors.append(f"第{row}行：专业 '{major_name}' 不存在")
                        continue
                    major_id = major.id

                # 学生角色验证
                if role == "student":
                    if not faculty_id or not department_id or not major_id:
                        error_count += 1
                        errors.append(f"第{row}行：学生用户必须填写学院、系和专业")
                        continue

                # 创建用户
                new_user = User(
                    username=username,
                    name=name,
                    role=role,
                    faculty_id=faculty_id,
                    department_id=department_id,
                    major_id=major_id,
                    email=email or "",
                    phone=phone or "",
                    status=status,
                )

                # 设置默认密码
                new_user.set_password("123456")

                # 学生角色需要创建学生记录
                if role == "student":
                    # 检查是否已存在对应的Student记录
                    existing_student = Student.query.filter_by(
                        student_id=username
                    ).first()
                    if not existing_student:
                        new_student = Student(
                            student_id=username,
                            student_name=name,
                            faculty_id=faculty_id,
                            department_id=department_id,
                            major_id=major_id,
                            gender=gender,
                            cet4_score=cet4_score,
                            cet6_score=cet6_score,
                            gpa=gpa,
                            academic_score=academic_score,
                        )
                        db.session.add(new_student)
                        new_user.student_id = username

                # 添加用户到数据库
                db.session.add(new_user)
                success_count += 1

            except Exception as e:
                error_count += 1
                errors.append(f"第{row}行：导入失败 - {str(e)}")

        # 提交事务
        db.session.commit()

        return (
            jsonify(
                {
                    "message": "用户导入完成",
                    "success_count": success_count,
                    "error_count": error_count,
                    "errors": errors,
                }
            ),
            200,
        )

    except Exception as e:
        db.session.rollback()
        return jsonify({"message": f"导入失败：{str(e)}"}), 500
